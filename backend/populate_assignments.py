import os
import re
import openpyxl
import psycopg2
from dotenv import load_dotenv
from datetime import datetime, timezone

# Load environment
backend_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(backend_dir, ".env"))

db_url = os.getenv("NEON_DB_URL") or os.getenv("DATABASE_URL")
if not db_url:
    print("Error: DATABASE_URL or NEON_DB_URL not set in environment.")
    exit(1)

excel_path = r"C:\Users\tejas\Downloads\5.Hackathon 2026 Evaluation_Panel.xlsx"
if not os.path.exists(excel_path):
    print(f"Error: {excel_path} not found.")
    exit(1)

print("Connecting to database...")
conn = psycopg2.connect(db_url)
cursor = conn.cursor()

# Ensure Desc Maker exists
cursor.execute("SELECT id FROM teams WHERE LOWER(TRIM(name)) = 'desc maker';")
desc_maker_exists = cursor.fetchone()
if not desc_maker_exists:
    print("Inserting missing team 'Desc Maker'...")
    created_at = datetime.now(timezone.utc).isoformat()
    cursor.execute(
        """
        INSERT INTO teams (name, registered, leader_email, disqualified, members, domain, venue, created_at)
        VALUES ('Desc Maker', 0, NULL, 0, '["Priyansh"]', 'Health Care', 'A Block & C Block', %s)
        """,
        (created_at,)
    )
    conn.commit()

# Load all teams into a dictionary (normalized_name -> id)
cursor.execute("SELECT id, name FROM teams;")
teams_db = {}
for tid, tname in cursor.fetchall():
    norm = tname.strip().lower().replace(" ", "")
    teams_db[norm] = tid

# Load all judges into a dictionary (normalized_name -> id)
cursor.execute("SELECT id, name FROM users WHERE role = 'judge';")
judges_db = {}
for jid, jname in cursor.fetchall():
    norm = jname.strip().lower().replace(" ", "")
    judges_db[norm] = jid

# Clear old judge assignments
print("Clearing existing judge assignments...")
cursor.execute("DELETE FROM assignments WHERE person_role = 'judge';")
conn.commit()

# Load workbook
wb = openpyxl.load_workbook(excel_path)

def normalize_judge_name(raw_name):
    if not raw_name:
        return ""
    name = raw_name.replace("Mr. ", "").strip().lower().replace(" ", "")
    # Special mapping for Mr. Nandkishore -> Nandkishore Patidar
    if name == "nandkishore":
        return "nandkishorepatidar"
    return name

assignment_count = 0
not_found_teams = []

for sheet_name in wb.sheetnames:
    if sheet_name == 'Winners':
        continue
    sheet = wb[sheet_name]
    
    # Read judges
    j1_raw = sheet.cell(3, 5).value
    j2_raw = sheet.cell(4, 5).value
    
    j1_norm = normalize_judge_name(j1_raw)
    j2_norm = normalize_judge_name(j2_raw)
    
    j1_id = judges_db.get(j1_norm)
    j2_id = judges_db.get(j2_norm)
    
    if not j1_id:
        print(f"Warning: Judge '{j1_raw}' (normalized: '{j1_norm}') not found in database.")
    if not j2_id:
        print(f"Warning: Judge '{j2_raw}' (normalized: '{j2_norm}') not found in database.")
        
    print(f"\nProcessing Sheet: {sheet_name} | Judges: {j1_raw} (ID: {j1_id}) and {j2_raw} (ID: {j2_id})")
    
    # Read teams
    for r in range(7, 100):
        t_name = sheet.cell(r, 2).value
        if not t_name:
            # If we hit an empty row, check if next rows are empty
            if all(sheet.cell(r+i, 2).value is None for i in range(1, 5)):
                break
            continue
            
        t_name_str = str(t_name).strip()
        t_norm = t_name_str.lower().replace(" ", "")
        
        t_id = teams_db.get(t_norm)
        if not t_id:
            # Try matching with minor punctuation changes
            t_norm_clean = re.sub(r"[^\w]", "", t_norm)
            matched = False
            for db_norm, db_id in teams_db.items():
                if re.sub(r"[^\w]", "", db_norm) == t_norm_clean:
                    t_id = db_id
                    matched = True
                    break
            if not matched:
                print(f"  Warning: Team '{t_name_str}' not found in database.")
                not_found_teams.append((sheet_name, t_name_str))
                continue
                
        # Insert assignments
        if j1_id:
            cursor.execute(
                "INSERT INTO assignments (person_role, person_id, team_id) VALUES ('judge', %s, %s) ON CONFLICT DO NOTHING",
                (j1_id, t_id)
            )
            assignment_count += 1
        if j2_id:
            cursor.execute(
                "INSERT INTO assignments (person_role, person_id, team_id) VALUES ('judge', %s, %s) ON CONFLICT DO NOTHING",
                (j2_id, t_id)
            )
            assignment_count += 1

conn.commit()
conn.close()

print(f"\nDone! Inserted {assignment_count} judge assignments.")
if not_found_teams:
    print(f"Failed to find {len(not_found_teams)} teams in DB:")
    for s_name, t_name in not_found_teams:
        print(f"  - Sheet: {s_name}, Team: {t_name}")
